from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from .models import BoardEscalation, DomainStats


# ---- CONFIG -------------------------------------------------------------

MASTER_SUMMARY_SHEET = "Master_Summary"
FUNDAMENTALS_SHEET = "Token Fundamentals & Governance"

# Sheets we do NOT scan for board escalations
IGNORED_SHEETS = {
    MASTER_SUMMARY_SHEET,
    "Overview",
}

# Sheets we scan for per-question responses (signal extraction). Usually these
# are the domain tabs, but we keep it generic: any sheet with Question_ID +
# Raw_Response is eligible.


# Header patterns (lowercase substring matches) for Master_Summary
DOMAIN_COL_HEADERS = ["domain"]
WEIGHT_COL_HEADERS = ["weight"]
AVG_SCORE_HEADERS = ["domain_avg_final_score", "average", "avg"]
BAND_HEADERS = ["domain_risk_band", "risk band"]

# Header patterns for domain sheets (board escalation data)
QUESTION_ID_HEADERS = ["question_id"]
QUESTION_TEXT_HEADERS = ["question_text"]
ESC_FLAG_HEADERS = ["board_escalation_flag"]
TRIGGER_RULE_HEADERS = ["trigger_rule_description"]
NARRATIVE_HEADERS = ["narrative_justification"]
CITATIONS_HEADERS = ["source_citations"]
SOURCE_DATE_HEADERS = ["most_recent_source_date"]
STALENESS_HEADERS = ["staleness_class"]

# Header patterns for token fundamentals
RAW_RESPONSE_HEADERS = ["raw_response", "raw response", "response"]
CONFIDENCE_HEADERS = ["confidence"]

# For generic response extraction we also want the following fields when present
RAW_POINTS_HEADERS = ["raw_points", "raw points"]
FINAL_SCORE_HEADERS = ["final_score", "final score"]


# ---- ESCALATION CLASSIFICATION -----------------------------------------

# Flags that mean "this is NOT a real board trigger"
NON_ESCALATION_FLAGS = {
    "",
    "no",
    "false",
    "0",
    "no review",  # informational only
}

# Substrings that indicate a real escalation
REAL_ESCALATION_KEYWORDS = [
    "review required",
    "board review",
    "listing committee",
    "escalate",
    "reject",
]


def _is_real_board_trigger(flag: str) -> bool:
    """
    Return True only for flags that mean 'this needs board attention',
    not for 'No Review' informational narratives.
    """
    if flag is None:
        return False
    f = flag.strip().lower()
    if f in NON_ESCALATION_FLAGS:
        return False
    return any(k in f for k in REAL_ESCALATION_KEYWORDS)


# ---- HEADER UTILS ------------------------------------------------------

def _normalise_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _find_header_row(ws, max_search_rows: int = 5) -> Optional[int]:
    """
    Try to find the header row by looking for something that looks like
    'question_id' or 'domain' in the first few rows.
    """
    for row in range(1, max_search_rows + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        norm_values = [_normalise_header(v) for v in values]
        if any("question_id" in v or "domain" in v for v in norm_values):
            return row
    return None


def _build_header_map(ws, header_row: int) -> Dict[str, int]:
    """
    Map normalised header text -> column index (1-based).
    """
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        norm = _normalise_header(value)
        if norm:
            header_map[norm] = col
    return header_map


def _find_first_matching_col(header_map: Dict[str, int], patterns: List[str]) -> Optional[int]:
    for header, col in header_map.items():
        for p in patterns:
            if p in header:
                return col
    return None


def _band_numeric_from_name(band_name: str) -> int:
    """
    Map band strings to numeric 1–5.
    """
    name = band_name.strip().lower()
    if name.startswith("very low"):
        return 1
    if name.startswith("low"):
        return 2
    if name.startswith("medium-high"):
        return 4
    if name.startswith("high"):
        return 5
    if name.startswith("medium"):
        return 3
    return 0


def _band_name_from_score(score: float) -> str:
    """
    Fallback if there is no band column.
    Mirrors your Excel formula:
      <3 Very Low, <6 Low, <9 Medium, <12 Medium-High, else High
    """
    if score < 3:
        return "Very Low"
    if score < 6:
        return "Low"
    if score < 9:
        return "Medium"
    if score < 12:
        return "Medium-High"
    return "High"


# ---- PARSE DOMAIN STATS ------------------------------------------------

def parse_domain_stats(wb) -> List[DomainStats]:
    """
    Parse domain-level stats from the Master_Summary sheet.

    Expected header row contains something like:
      Domain | Weight | Domain_Avg_Final_Score | Domain_Risk_Band | ...
    """
    if MASTER_SUMMARY_SHEET not in wb.sheetnames:
        raise ValueError(f"Expected sheet '{MASTER_SUMMARY_SHEET}' not found")

    ws = wb[MASTER_SUMMARY_SHEET]

    header_row = _find_header_row(ws)
    if header_row is None:
        raise ValueError(f"Could not find header row in '{MASTER_SUMMARY_SHEET}'")

    header_map = _build_header_map(ws, header_row)

    domain_col = _find_first_matching_col(header_map, DOMAIN_COL_HEADERS)
    if not domain_col:
        raise ValueError("Could not find a 'Domain' column in Master_Summary")

    weight_col = _find_first_matching_col(header_map, WEIGHT_COL_HEADERS)
    avg_score_col = _find_first_matching_col(header_map, AVG_SCORE_HEADERS)
    band_col = _find_first_matching_col(header_map, BAND_HEADERS)

    domain_stats: List[DomainStats] = []

    row = header_row + 1
    while row <= ws.max_row:
        domain_name = ws.cell(row=row, column=domain_col).value
        if domain_name is None or str(domain_name).strip() == "":
            # Stop when we hit a blank domain row
            break

        domain_name_str = str(domain_name).strip()

        weight_val = ws.cell(row=row, column=weight_col).value if weight_col else None
        avg_score_val = ws.cell(row=row, column=avg_score_col).value if avg_score_col else None
        band_name_val = ws.cell(row=row, column=band_col).value if band_col else None

        try:
            weight = float(weight_val) if weight_val is not None else 0.0
        except (TypeError, ValueError):
            weight = 0.0

        try:
            avg_score = float(avg_score_val) if avg_score_val is not None else 0.0
        except (TypeError, ValueError):
            avg_score = 0.0

        if band_name_val:
            band_name = str(band_name_val).strip()
        else:
            band_name = _band_name_from_score(avg_score)

        band_numeric = _band_numeric_from_name(band_name)

        stats = DomainStats(
            code=domain_name_str,       # no separate domain code column, so use the name for now
            name=domain_name_str,
            weight=weight,
            avg_score=avg_score,
            band_name=band_name,
            band_numeric=band_numeric,
            has_board_escalation=False,   # will be populated later
            board_escalation_count=0,     # will be populated later
        )
        domain_stats.append(stats)
        row += 1

    return domain_stats


# ---- PARSE BOARD ESCALATIONS -------------------------------------------

def parse_board_escalations(wb) -> List[BoardEscalation]:
    """
    Scan all non-ignored sheets for rows where a 'Board_Escalation_Flag' is set.

    Uses headers so you can keep changing the underlying Excel logic.
    We keep *all* rows with any flag (including 'No Review'), then later decide
    which ones are “real” board triggers via _is_real_board_trigger().
    """
    escalations: List[BoardEscalation] = []

    for sheet_name in wb.sheetnames:
        if sheet_name in IGNORED_SHEETS:
            continue

        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        if header_row is None:
            continue

        header_map = _build_header_map(ws, header_row)

        qid_col = _find_first_matching_col(header_map, QUESTION_ID_HEADERS)
        qtext_col = _find_first_matching_col(header_map, QUESTION_TEXT_HEADERS)
        esc_flag_col = _find_first_matching_col(header_map, ESC_FLAG_HEADERS)
        trigger_col = _find_first_matching_col(header_map, TRIGGER_RULE_HEADERS)
        narrative_col = _find_first_matching_col(header_map, NARRATIVE_HEADERS)
        citations_col = _find_first_matching_col(header_map, CITATIONS_HEADERS)
        date_col = _find_first_matching_col(header_map, SOURCE_DATE_HEADERS)
        stale_col = _find_first_matching_col(header_map, STALENESS_HEADERS)

        if not esc_flag_col:
            # No escalation flag column on this sheet – skip
            continue

        domain_name = sheet_name
        domain_code = sheet_name  # you can map to short codes later if you want

        row = header_row + 1
        while row <= ws.max_row:
            qid_val = ws.cell(row=row, column=qid_col).value if qid_col else None
            qtext_val = ws.cell(row=row, column=qtext_col).value if qtext_col else None
            esc_flag_val = ws.cell(row=row, column=esc_flag_col).value

            # if both ID and text are blank, just move on
            if (qid_val is None or str(qid_val).strip() == "") and (
                qtext_val is None or str(qtext_val).strip() == ""
            ):
                row += 1
                continue

            flag_str = str(esc_flag_val or "").strip()

            # If there's no flag at all, skip this row
            if flag_str == "":
                row += 1
                continue

            question_id = str(qid_val or "").strip()
            question_text = str(qtext_val or "").strip()

            trigger_rule_val = ws.cell(row=row, column=trigger_col).value if trigger_col else None
            narrative_val = ws.cell(row=row, column=narrative_col).value if narrative_col else None
            citations_val = ws.cell(row=row, column=citations_col).value if citations_col else None
            date_val = ws.cell(row=row, column=date_col).value if date_col else None
            stale_val = ws.cell(row=row, column=stale_col).value if stale_col else None

            trigger_rule = str(trigger_rule_val).strip() if trigger_rule_val else None
            raw_narrative = str(narrative_val).strip() if narrative_val else None
            staleness_class = str(stale_val).strip() if stale_val else None

            citations: List[Dict[str, str]] = []
            if citations_val:
                # Assuming semi-colon separated URLs/labels for now
                parts = str(citations_val).split(";")
                for p in parts:
                    p = p.strip()
                    if p:
                        citations.append({"label": p, "url": p})

            most_recent_source_date = None
            if date_val:
                most_recent_source_date = str(date_val)

            escalation_id = f"{domain_code}_{question_id}" if question_id else f"{domain_code}_{row}"

            esc = BoardEscalation(
                id=escalation_id,
                domain_code=domain_code,
                domain_name=domain_name,
                question_id=question_id,
                question_text=question_text,
                flag=flag_str,
                trigger_rule=trigger_rule,
                raw_narrative=raw_narrative,
                most_recent_source_date=most_recent_source_date,
                staleness_class=staleness_class,
                citations=citations,
            )
            escalations.append(esc)

            row += 1

    return escalations


# ---- PARSE PER-QUESTION RESPONSES (FOR SIGNALS) ------------------------

def parse_question_responses(wb) -> Dict[str, Any]:
    """Extract all per-question responses from eligible DDQ tabs.

    This is used by the deterministic tag inference layer (signals + gating).

    Output:
      {
        "responses": List[Dict[str, Any]],
        "answers_by_key": Dict[str, List[Dict[str, Any]]],  # keyed by "<Sheet>::<QID>"
      }
    """

    responses: List[Dict[str, Any]] = []
    answers_by_key: Dict[str, List[Dict[str, Any]]] = {}

    def key(sheet: str, qid: str) -> str:
        return f"{sheet}::{str(qid or '').strip().upper()}"

    for sheet_name in wb.sheetnames:
        if sheet_name in IGNORED_SHEETS:
            continue

        ws = wb[sheet_name]
        header_row = _find_header_row(ws)
        if header_row is None:
            continue

        header_map = _build_header_map(ws, header_row)

        qid_col = _find_first_matching_col(header_map, QUESTION_ID_HEADERS)
        qtext_col = _find_first_matching_col(header_map, QUESTION_TEXT_HEADERS)
        raw_col = _find_first_matching_col(header_map, RAW_RESPONSE_HEADERS)
        conf_col = _find_first_matching_col(header_map, CONFIDENCE_HEADERS)
        narrative_col = _find_first_matching_col(header_map, NARRATIVE_HEADERS)
        citations_col = _find_first_matching_col(header_map, CITATIONS_HEADERS)
        esc_flag_col = _find_first_matching_col(header_map, ESC_FLAG_HEADERS)
        trigger_col = _find_first_matching_col(header_map, TRIGGER_RULE_HEADERS)
        raw_points_col = _find_first_matching_col(header_map, RAW_POINTS_HEADERS)
        final_score_col = _find_first_matching_col(header_map, FINAL_SCORE_HEADERS)

        # Require at least Question_ID and Raw_Response to treat this as an answer sheet.
        if not qid_col or not raw_col:
            continue

        for row in range(header_row + 1, ws.max_row + 1):
            qid_val = ws.cell(row=row, column=qid_col).value
            if qid_val is None or str(qid_val).strip() == "":
                continue

            qid = str(qid_val).strip()
            qtext_val = ws.cell(row=row, column=qtext_col).value if qtext_col else None
            raw_val = ws.cell(row=row, column=raw_col).value

            # Skip section header rows (some templates have a label row with QID=None,
            # but we already handled that; keep this extra guard for weird inputs).
            if qid.lower() in {"none", "nan"}:
                continue

            conf_val = ws.cell(row=row, column=conf_col).value if conf_col else None
            narrative_val = ws.cell(row=row, column=narrative_col).value if narrative_col else None
            citations_val = ws.cell(row=row, column=citations_col).value if citations_col else None
            esc_flag_val = ws.cell(row=row, column=esc_flag_col).value if esc_flag_col else None
            trigger_val = ws.cell(row=row, column=trigger_col).value if trigger_col else None
            raw_points_val = ws.cell(row=row, column=raw_points_col).value if raw_points_col else None
            final_score_val = ws.cell(row=row, column=final_score_col).value if final_score_col else None

            citations: List[str] = []
            if citations_val:
                for p in str(citations_val).split(";"):
                    p = p.strip()
                    if p:
                        citations.append(p)

            rec: Dict[str, Any] = {
                "sheet": sheet_name,
                "question_id": qid,
                "question_text": str(qtext_val or "").strip() if qtext_val else "",
                "raw_response": str(raw_val or "").strip() if raw_val is not None else "",
                "confidence": str(conf_val or "").strip() if conf_val is not None else "",
                "narrative_justification": str(narrative_val or "").strip() if narrative_val is not None else "",
                "source_citations": citations,
                "board_escalation_flag": str(esc_flag_val or "").strip() if esc_flag_val is not None else "",
                "trigger_rule_description": str(trigger_val or "").strip() if trigger_val is not None else "",
                "raw_points": raw_points_val,
                "final_score": final_score_val,
                "row_number": row,
            }
            responses.append(rec)
            answers_by_key.setdefault(key(sheet_name, qid), []).append(rec)

    return {
        "responses": responses,
        "answers_by_key": answers_by_key,
    }


# ---- PARSE TOKEN CATEGORY (A1.1) --------------------------------------

def _parse_primary_secondary(raw: str) -> Dict[str, Optional[str]]:
    """Parse strings like 'Primary: Native L1; Secondary: Gas/Fee'."""
    if not raw:
        return {"primary": None, "secondary": None}

    txt = str(raw).strip()
    primary = None
    secondary = None

    # Allow separators like ';' or ','
    parts = [p.strip() for p in txt.replace(",", ";").split(";") if p and p.strip()]
    for p in parts:
        low = p.lower()
        if low.startswith("primary"):
            val = p.split(":", 1)[1] if ":" in p else p
            primary = val.strip() or None
        elif low.startswith("secondary"):
            val = p.split(":", 1)[1] if ":" in p else p
            secondary = val.strip() or None

    # Fallback: if no labels detected but we have 2 parts
    if primary is None and secondary is None and len(parts) == 2:
        primary, secondary = parts[0], parts[1]

    return {"primary": primary, "secondary": secondary}


def parse_token_category(wb) -> Optional[Dict[str, Any]]:
    """Extract A1.1 from the 'Token Fundamentals & Governance' sheet."""
    if FUNDAMENTALS_SHEET not in wb.sheetnames:
        return None

    ws = wb[FUNDAMENTALS_SHEET]
    header_row = _find_header_row(ws)
    if header_row is None:
        return None

    header_map = _build_header_map(ws, header_row)
    qid_col = _find_first_matching_col(header_map, QUESTION_ID_HEADERS)
    raw_col = _find_first_matching_col(header_map, RAW_RESPONSE_HEADERS)
    conf_col = _find_first_matching_col(header_map, CONFIDENCE_HEADERS)
    narrative_col = _find_first_matching_col(header_map, NARRATIVE_HEADERS)

    if not qid_col or not raw_col:
        return None

    row = header_row + 1
    while row <= ws.max_row:
        qid_val = ws.cell(row=row, column=qid_col).value
        if qid_val is None:
            row += 1
            continue
        if str(qid_val).strip().upper() == "A1.1":
            raw_val = ws.cell(row=row, column=raw_col).value
            conf_val = ws.cell(row=row, column=conf_col).value if conf_col else None
            nar_val = ws.cell(row=row, column=narrative_col).value if narrative_col else None

            parsed = _parse_primary_secondary(str(raw_val or ""))
            return {
                "question_id": "A1.1",
                "raw": str(raw_val or "").strip() or None,
                "primary": parsed.get("primary"),
                "secondary": parsed.get("secondary"),
                "confidence": str(conf_val).strip() if conf_val else None,
                "narrative": str(nar_val).strip() if nar_val else None,
                "source_sheet": FUNDAMENTALS_SHEET,
            }
        row += 1

    return None


# ---- TOP-LEVEL ENTRY POINT ---------------------------------------------

def parse_ddq(ddq_path: Path) -> Dict[str, Any]:
    """
    High-level entry point for the rest of the app.

    Returns a dict with:
      - domain_stats: List[DomainStats]
      - board_escalations: List[BoardEscalation]
      - snapshot: {}                # placeholder for now
      - citations: []               # placeholder for now
      - project_description: ""     # placeholder for now
    """
    ddq_path = Path(ddq_path)
    if not ddq_path.exists():
        raise FileNotFoundError(f"DDQ workbook not found: {ddq_path}")

    wb = load_workbook(ddq_path, data_only=True)

    domain_stats = parse_domain_stats(wb)
    board_escalations = parse_board_escalations(wb)
    token_category = parse_token_category(wb)
    response_pack = parse_question_responses(wb)

    # Compute real board triggers per domain based on row-level flags
    counts = Counter()
    for esc in board_escalations:
        if _is_real_board_trigger(esc.flag):
            counts[esc.domain_name] += 1

    # Enrich domain_stats with counts and booleans
    for d in domain_stats:
        c = counts.get(d.name, 0)
        d.board_escalation_count = c
        d.has_board_escalation = c > 0

    snapshot: Dict[str, Any] = {}
    citations: List[Dict[str, str]] = []
    project_description = ""

    return {
        "domain_stats": domain_stats,
        "board_escalations": board_escalations,
        "token_category": token_category,
        "responses": response_pack["responses"],
        "answers_by_key": response_pack["answers_by_key"],
        "snapshot": snapshot,
        "citations": citations,
        "project_description": project_description,
    }
